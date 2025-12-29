"""
merge_to_json.py
A utility with GUI and CLI to merge multiple files (.txt/.csv/.xls/.xlsx/.parquet/.feather) into one output file.

Usage (GUI):
    python merge_to_json.py

Usage (CLI):
    python merge_to_json.py --cli file1.parquet file2.xlsx file3.csv --header 1 --format parquet-snappy -o merged.parquet

Output: Supported formats are JSON, NDJSON, Parquet, Feather, CSV, CSV+GZIP (utf-8, pretty printed)

Dependencies: pandas, openpyxl, xlrd, pyarrow (for parquet/feather)
"""

import os
import sys
import argparse
import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd


def read_file(filepath, header_row=1):
    """Read a file into a pandas DataFrame. For Excel files all sheets are returned as dict of DataFrames.
    For parquet/feather returns a single DataFrame.
    For txt/csv returns a single DataFrame.
    header_row is 1-based index of header line.
    """
    ext = os.path.splitext(filepath)[1].lower()
    
    # Parquet format
    if ext == ".parquet":
        try:
            df = pd.read_parquet(filepath)
            return df  # return single DataFrame (not dict)
        except Exception as e:
            raise RuntimeError(f"Gagal membaca file parquet {filepath}: {e}")
    
    # Feather format
    elif ext == ".feather":
        try:
            df = pd.read_feather(filepath)
            return df  # return single DataFrame
        except Exception as e:
            raise RuntimeError(f"Gagal membaca file feather {filepath}: {e}")
    
    # Excel format
    elif ext in [".xlsx", ".xls"]:
        # read all sheets
        try:
            sheets = pd.read_excel(filepath, sheet_name=None, engine="pyarrow" if ext == ".xlsx" else None, header=header_row-1)
        except Exception:
            # try openpyxl for xlsx
            try:
                sheets = pd.read_excel(filepath, sheet_name=None, engine="openpyxl", header=header_row-1)
            except Exception:
                # try xlrd for xls
                sheets = pd.read_excel(filepath, sheet_name=None, header=header_row-1)
        return sheets  # dict: sheet_name -> df
    
    # CSV/TXT format
    else:
        # try csv/txt delimiter detection
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            sample = f.read(4096)
        if "\t" in sample:
            delimiter = "\t"
        elif ";" in sample:
            delimiter = ";"
        else:
            delimiter = ","
        df = pd.read_csv(filepath, delimiter=delimiter, header=header_row-1, encoding="utf-8", engine="python")
        return df


def merge_files(filepaths, header_row=1, include_source=True, include_sheet=True, sheet_selection=None, progress_callback=None):
    """Merge multiple files and return a single DataFrame.

    sheet_selection: optional dict {filepath: [sheet1, sheet2, ...]} to limit sheets per Excel file.
    progress_callback: optional callable(current, total) to report progress.
    """
    dfs = []

    # Compute total steps for progress: for non-excel -> 1, for excel -> number of sheets to read
    total_steps = 0
    for fp in filepaths:
        ext = os.path.splitext(fp)[1].lower()
        if ext in [".xlsx", ".xls"]:
            try:
                xl = pd.ExcelFile(fp, engine="openpyxl" if fp.lower().endswith('.xlsx') else None)
                sheets = xl.sheet_names
            except Exception:
                try:
                    xl = pd.ExcelFile(fp)
                    sheets = xl.sheet_names
                except Exception:
                    sheets = []
            if sheet_selection and fp in sheet_selection and sheet_selection[fp]:
                total_steps += len(sheet_selection[fp])
            else:
                total_steps += max(1, len(sheets))
        else:
            total_steps += 1

    step = 0
    for fp in filepaths:
        ext = os.path.splitext(fp)[1].lower()
        try:
            if ext in [".xlsx", ".xls"]:
                # Determine which sheets to read
                sel = None
                if sheet_selection and fp in sheet_selection and sheet_selection[fp]:
                    sel = sheet_selection[fp]
                if sel is None:
                    # read all sheets
                    try:
                        res = pd.read_excel(fp, sheet_name=None, engine="pyarrow", header=header_row-1)
                    except Exception:
                        res = pd.read_excel(fp, sheet_name=None, header=header_row-1)
                else:
                    try:
                        res = pd.read_excel(fp, sheet_name=sel, engine="pyarrow", header=header_row-1)
                    except Exception:
                        res = pd.read_excel(fp, sheet_name=sel, header=header_row-1)
                # res can be dict or DataFrame (if single sheet requested)
                if isinstance(res, dict):
                    items = res.items()
                else:
                    # single sheet -> try to get name from selection or use generic
                    name = sel[0] if isinstance(sel, (list, tuple)) and sel else os.path.splitext(os.path.basename(fp))[0]
                    items = [(name, res)]
                for sheet_name, df in items:
                    df = df.copy()
                    if include_source:
                        df["Sumber File"] = os.path.basename(fp)
                    if include_sheet:
                        df["Sumber Sheet"] = sheet_name
                    dfs.append(df)
                    step += 1
                    if progress_callback:
                        progress_callback(step, total_steps)
            else:
                res = read_file(fp, header_row=header_row)
                df = res.copy()
                if include_source:
                    df["Sumber File"] = os.path.basename(fp)
                if include_sheet:
                    df["Sumber Sheet"] = ""
                dfs.append(df)
                step += 1
                if progress_callback:
                    progress_callback(step, total_steps)
        except Exception as e:
            raise RuntimeError(f"Gagal membaca {fp}: {e}")

    if not dfs:
        return pd.DataFrame()
    # Align columns: union of all columns, fill NaN where missing
    all_cols = []
    for d in dfs:
        for c in d.columns:
            if c not in all_cols:
                all_cols.append(c)
    normalized = []
    for d in dfs:
        missing = [c for c in all_cols if c not in d.columns]
        if missing:
            for m in missing:
                d[m] = pd.NA
        normalized.append(d[all_cols])
    result = pd.concat(normalized, ignore_index=True)
    
    # Pastikan SEMUA kolom yang memiliki leading zeros tidak hilang.
    # Scan semua kolom dan jika ditemukan nilai dengan leading zero, 
    # konversi kolom tersebut menjadi string dtype untuk preserve leading zeros.
    for col in result.columns:
        non_null = result[col].dropna()
        if len(non_null) == 0:
            continue
        
        # Cek apakah ada leading zeros di kolom ini
        has_leading_zero = False
        for v in non_null:
            s = str(v).strip()
            # Jika string numeric dengan leading zero
            if s.isdigit() and len(s) > 1 and s[0] == '0':
                has_leading_zero = True
                break
        
        # Jika ada leading zero, convert ke string dtype untuk preserve
        if has_leading_zero:
            try:
                result[col] = result[col].astype(str)
            except Exception:
                # Jika gagal, gunakan apply sebagai fallback
                result[col] = result[col].apply(lambda x: str(x) if pd.notna(x) else x)
            print(f"Kolom '{col}': Converted to string to preserve leading zeros")
    
    # Keep existing MRN-specific logic untuk backward compatibility
    for col in result.columns:
        col_str = str(col).strip()
        # Force known MRN column name to string
        if col_str.lower() == 'mrn':
            # Ensure MRN values are 6-digit strings with leading zeros when needed.
            def pad_mrn(val):
                if pd.isna(val):
                    return pd.NA
                s = str(val).strip()
                if s == '':
                    return pd.NA
                # If it's numeric-like (int/float or digit-only string), zero-pad to 6 digits
                try:
                    # handle floats like 1234.0
                    if isinstance(val, (int,)):
                        return str(val).zfill(6)
                    if isinstance(val, float):
                        if val.is_integer():
                            return str(int(val)).zfill(6)
                    if s.isdigit():
                        return s.zfill(6)
                    # try to parse numeric from string (e.g., '1234.0')
                    if '.' in s:
                        f = float(s)
                        if f.is_integer():
                            return str(int(f)).zfill(6)
                except Exception:
                    pass
                # Non-numeric strings: return trimmed string unchanged
                return s

            try:
                result[col] = result[col].apply(lambda x: None if pd.isna(x) else pad_mrn(x)).astype('string')
            except Exception:
                result[col] = result[col].apply(lambda x: None if pd.isna(x) else pad_mrn(x))
            continue
    
    # Jangan modifikasi atau parsing kolom lainnya - pertahankan sesuai file asli
    return result


def standardize_date_columns(df):
    """Standardize all date columns to dd/mm/yyyy string format.
    Detect columns by name (containing date/tanggal/tgl/waktu/time) or datetime dtype.
    Also detects and converts Excel serial numbers (e.g., 45883) in any format.
    Returns a copy of df with date columns standardized.
    """
    df_std = df.copy()
    
    for col in df_std.columns:
        col_name_lower = str(col).lower().strip()
        is_date_col = any(k in col_name_lower for k in ('date', 'tanggal', 'tgl', 'waktu', 'time', 'lahir'))
        is_datetime_dtype = pd.api.types.is_datetime64_any_dtype(df_std[col])
        
        if is_date_col or is_datetime_dtype:
            # Convert to dd/mm/yyyy string format
            try:
                # Strategy 1: Try convert Excel serial numbers (numeric or string that looks numeric)
                # Excel serial numbers are typically 5-digit numbers (1-100000)
                try:
                    # Convert column to numeric, coerce errors
                    numeric_col = pd.to_numeric(df_std[col], errors='coerce')
                    # Check if most values are valid numbers in Excel date range (1 to 60000)
                    valid_numeric = (numeric_col.notna()) & (numeric_col > 0) & (numeric_col < 100000)
                    valid_count = valid_numeric.sum()
                    if valid_count / max(1, len(numeric_col)) >= 0.5:
                        # Likely Excel serial numbers, convert
                        converted = pd.to_datetime(numeric_col, unit='D', origin=pd.Timestamp('1899-12-30'), errors='coerce')
                        df_std[col] = converted
                        is_datetime_dtype = True
                except Exception:
                    pass
                
                # Strategy 2: Try general datetime parsing if not already converted
                if not is_datetime_dtype:
                    try:
                        parsed = pd.to_datetime(df_std[col], errors='coerce', dayfirst=True)
                        valid_count = parsed.notna().sum()
                        # Check how many are epoch (1970-01-01)
                        epoch_count = 0
                        if valid_count > 0:
                            epoch_count = ((parsed.dt.year == 1970) & (parsed.dt.month == 1) & (parsed.dt.day == 1)).sum()
                        
                        # Only use parsed if most succeeded and not too many epochs
                        if valid_count / max(1, len(parsed)) >= 0.5 and epoch_count / max(1, len(parsed)) < 0.2:
                            df_std[col] = parsed
                    except Exception:
                        pass
                
                # Format as dd/mm/yyyy
                def format_date(v):
                    if pd.isna(v):
                        return None
                    try:
                        # Handle various date types
                        if isinstance(v, (pd.Timestamp, datetime.datetime)):
                            # Check if epoch (1970-01-01)
                            if v.year == 1970 and v.month == 1 and v.day == 1:
                                return str(v)
                            return v.strftime('%d/%m/%Y')
                        elif isinstance(v, datetime.date):
                            if v.year == 1970 and v.month == 1 and v.day == 1:
                                return str(v)
                            return v.strftime('%d/%m/%Y')
                        else:
                            # Try parse string as datetime
                            dt = pd.to_datetime(v, errors='coerce', dayfirst=True)
                            if pd.notna(dt):
                                if dt.year == 1970 and dt.month == 1 and dt.day == 1:
                                    return str(v)
                                return dt.strftime('%d/%m/%Y')
                            else:
                                return str(v)
                    except Exception:
                        return str(v)
                
                df_std[col] = df_std[col].apply(format_date)
            except Exception:
                # If formatting fails, just convert to string
                df_std[col] = df_std[col].astype(str)
    
    return df_std


def save_to_json(df, outpath):
    """Save DataFrame to JSON file (list of records)."""
    try:
        import json, math
        import numpy as np
        import datetime as _dt

        def sanitize_value(v):
            # Convert pandas NA / NaN to None
            try:
                if pd.isna(v):
                    return None
            except Exception:
                pass

            # None stays None
            if v is None:
                return None

            # pandas Timestamp / datetime -> convert to string as-is
            if isinstance(v, (pd.Timestamp, _dt.datetime, _dt.date)):
                try:
                    return str(v)
                except Exception:
                    return None

            # numpy scalar -> native python types
            if isinstance(v, (np.integer,)):
                return int(v)
            if isinstance(v, (np.floating,)):
                f = float(v)
                if math.isfinite(f):
                    return f
                return str(v)
            if isinstance(v, (np.bool_,)):
                return bool(v)
            if isinstance(v, (np.ndarray,)):
                return [sanitize_value(x) for x in v.tolist()]

            # lists / tuples -> sanitize elements
            if isinstance(v, (list, tuple)):
                return [sanitize_value(x) for x in v]

            # dict -> sanitize recursively
            if isinstance(v, dict):
                return {str(k): sanitize_value(val) for k, val in v.items()}

            # Try native json dump; fallback to string or integer
            try:
                json.dumps(v)
                return v
            except (TypeError, OverflowError):
                # Try convert to integer first (but skip if looks like date or has special chars)
                try:
                    # Jika bisa convert ke int (misal dari float atau string angka), gunakan int
                    if isinstance(v, (int, float, str)):
                        str_v = str(v).strip()
                        # Jangan convert jika value adalah 'nan', 'inf' atau sejenisnya
                        if str_v.lower() in ('nan', 'inf', '-inf', 'infinity', '-infinity', 'none', ''):
                            raise ValueError("Invalid numeric value")
                        # Jangan convert jika terlihat seperti tanggal (ada / atau - atau :)
                        if any(c in str_v for c in ('/', '-', ':')):
                            raise ValueError("Looks like date/time, keep as string")
                        # Convert via float dulu (untuk handle scientific notation), lalu ke int
                        float_v = float(str_v)
                        if math.isfinite(float_v):
                            int_val = int(float_v)
                            return int_val
                except (ValueError, TypeError, OverflowError):
                    pass
                # Fallback ke string
                try:
                    return str(v)
                except Exception:
                    return None

        # Build records and sanitize all values so JSON can always be produced
        records = df.where(pd.notnull(df), None).to_dict(orient="records")
        sanitized = []
        for rec in records:
            sanitized.append({str(k): sanitize_value(v) for k, v in rec.items()})

        if outpath.lower().endswith('.ndjson'):
            # write newline-delimited JSON
            with open(outpath, "w", encoding="utf-8") as f:
                for rec in sanitized:
                    f.write(json.dumps(rec, ensure_ascii=False))
                    f.write('\n')
        else:
            with open(outpath, "w", encoding="utf-8") as f:
                json.dump(sanitized, f, ensure_ascii=False, indent=2)
    except Exception as e:
        raise RuntimeError(f"Gagal menyimpan JSON: {e}")


def save_dataframe(df, outpath, fmt='json'):
    """Save DataFrame to various formats.

    fmt values: 'json', 'ndjson', 'parquet-snappy', 'parquet-gzip', 'feather', 'csv-gzip', 'csv'
    """
    try:
        # Standardize date columns to dd/mm/yyyy format across all formats
        df_processed = standardize_date_columns(df)
        
        f = fmt.lower()
        if f == 'json' or f == 'json-pretty':
            save_to_json(df_processed, outpath)
        elif f == 'ndjson' or outpath.lower().endswith('.ndjson'):
            # NDJSON
            save_to_json(df_processed, outpath)
        elif f in ('parquet-snappy', 'parquet-gzip', 'parquet'):
            # write parquet with pyarrow
            try:
                import pyarrow  # noqa: F401
            except Exception:
                raise RuntimeError('pyarrow required to write parquet. Install with pip install pyarrow')
            
            # Prepare DataFrame: convert non-numeric columns to string to avoid dtype inference issues
            df_prep = df_processed.copy()
            for col in df_prep.columns:
                dtype = df_prep[col].dtype
                # Keep numeric columns as-is
                if pd.api.types.is_numeric_dtype(dtype):
                    continue
                # Convert everything else to string (safer for mixed-type columns)
                try:
                    df_prep[col] = df_prep[col].astype(str)
                except Exception:
                    pass
            
            comp = 'snappy' if 'snappy' in f else ('gzip' if 'gzip' in f else 'snappy')
            df_prep.to_parquet(outpath, index=False, compression=comp, coerce_timestamps='ms')
        elif f == 'feather':
            try:
                import pyarrow  # feather uses pyarrow
            except Exception:
                raise RuntimeError('pyarrow required to write feather. Install with pip install pyarrow')
            df_processed.reset_index(drop=True).to_feather(outpath)
        elif f == 'csv-gzip' or outpath.lower().endswith('.csv.gz'):
            # write gzipped csv
            if not outpath.lower().endswith('.gz'):
                # if user explicitly requested csv-gzip, add .gz extension
                if f == 'csv-gzip':
                    outpath = outpath + '.gz'
            df_processed.to_csv(outpath, index=False, compression='gzip', encoding='utf-8')
        elif f == 'csv' or outpath.lower().endswith('.csv'):
            # plain CSV (no compression)
            df_processed.to_csv(outpath, index=False, compression=None, encoding='utf-8')
        elif f == 'xlsx' or outpath.lower().endswith('.xlsx'):
            # Write to Excel format
            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter
            except Exception:
                pass
            
            # Write to Excel
            df_processed.to_excel(outpath, index=False, sheet_name='Data', engine='openpyxl')
            
            # Optional: Format Excel file (header styling)
            try:
                from openpyxl import load_workbook
                wb = load_workbook(outpath)
                ws = wb.active
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for col_num, col_title in enumerate(df_processed.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Auto-adjust column widths
                for col_num, col_title in enumerate(df_processed.columns, 1):
                    max_length = max(
                        df_processed[col_title].astype(str).apply(len).max(),
                        len(str(col_title))
                    ) + 2
                    col_letter = get_column_letter(col_num)
                    ws.column_dimensions[col_letter].width = min(max_length, 50)
                
                wb.save(outpath)
            except Exception:
                # If styling fails, just use basic Excel without formatting
                pass
        else:
            # fallback to JSON
            save_to_json(df_processed, outpath)
    except Exception as e:
        raise RuntimeError(f"Gagal menyimpan file: {e}")


# --------- GUI ----------

def configure_sheets_dialog(parent, filepaths_getter, sheet_selection):
    """Opens a dialog to configure which sheets to read per Excel file.
    filepaths_getter: callable returning current list of filepaths
    sheet_selection: dict to be mutated with selections
    """
    fps = [p for p in filepaths_getter()]
    if not fps:
        messagebox.showwarning("Peringatan", "Belum ada file yang dipilih.")
        return

    dlg = tk.Toplevel(parent)
    dlg.title("Konfigurasi Sheet per-file")
    dlg.geometry("600x400")
    frame = ttk.Frame(dlg, padding=8)
    frame.pack(fill=tk.BOTH, expand=True)

    canv = tk.Canvas(frame)
    vsb = ttk.Scrollbar(frame, orient='vertical', command=canv.yview)
    inner = ttk.Frame(canv)
    inner_id = canv.create_window((0,0), window=inner, anchor='nw')
    canv.configure(yscrollcommand=vsb.set)
    canv.pack(side='left', fill='both', expand=True)
    vsb.pack(side='right', fill='y')

    def on_configure(e=None):
        canv.configure(scrollregion=canv.bbox('all'))
    inner.bind('<Configure>', on_configure)

    listboxes = {}

    for fp in fps:
        ext = os.path.splitext(fp)[1].lower()
        lbl = ttk.Label(inner, text=os.path.basename(fp))
        lbl.pack(anchor='w', pady=(8,0))
        if ext in ['.xlsx', '.xls']:
            try:
                xl = pd.ExcelFile(fp)
                sheets = xl.sheet_names
            except Exception:
                sheets = []
            lb = tk.Listbox(inner, selectmode='multiple', height=min(8, max(3, len(sheets))))
            for s in sheets:
                lb.insert(tk.END, s)
            lb.pack(fill='x')
            # preselect previous
            prev = sheet_selection.get(fp)
            if prev:
                for i, s in enumerate(sheets):
                    if s in prev:
                        lb.selection_set(i)
            listboxes[fp] = (lb, sheets)
            btn_all = ttk.Button(inner, text='Pilih Semua', command=lambda l=lb: (l.selection_set(0, tk.END)))
            btn_all.pack(anchor='e', pady=(2,4))
        else:
            ttk.Label(inner, text='(bukan Excel)').pack(anchor='w')

    def do_save():
        for fp, val in listboxes.items():
            lb, sheets = val
            sel = [sheets[i] for i in lb.curselection()]
            if sel:
                sheet_selection[fp] = sel
            else:
                # empty means all sheets
                sheet_selection.pop(fp, None)
        dlg.destroy()

    btn_frame = ttk.Frame(dlg)
    btn_frame.pack(fill='x', pady=6)
    ttk.Button(btn_frame, text='Simpan', command=do_save).pack(side='left', padx=6)
    ttk.Button(btn_frame, text='Batal', command=dlg.destroy).pack(side='left')


def run_gui():
    root = tk.Tk()
    root.title("Gabung File -> DDUAAR")
    root.geometry("700x420")

    files_var = tk.StringVar(value="")
    # sheet_selection: dict filepath -> list of sheet names (or empty list => all)
    sheet_selection = {}
    format_var = tk.StringVar(value="json")

    def pick_files():
        paths = filedialog.askopenfilenames(title="Pilih file (xlsx/xls/csv/txt/parquet/feather)", filetypes=[("All supported", "*.txt *.csv *.xlsx *.xls *.parquet *.feather"), ("Excel files", "*.xlsx *.xls"), ("Text files", "*.txt *.csv"), ("Parquet files", "*.parquet"), ("Feather files", "*.feather")])
        if paths:
            # limit to 20 for safety
            files_var.set('\n'.join(paths))

    def filepaths_getter():
        raw = files_var.get().strip()
        return raw.split('\n') if raw else []

    def do_merge():
        raw = files_var.get().strip()
        if not raw:
            messagebox.showwarning("Peringatan", "Pilih minimal 1 file terlebih dahulu.")
            return
        filepaths = raw.split('\n')
        if len(filepaths) > 20:
            messagebox.showwarning("Peringatan", "Pilih maksimal 20 file pada GUI (gunakan CLI untuk lebih).")
            return
        try:
            header_row = int(header_entry.get())
            if header_row < 1:
                header_row = 1
        except Exception:
            header_row = 1
        include_source = source_var.get()
        include_sheet = sheet_var.get()
        try:
            # Disable controls while merging
            btn_merge.config(state='disabled')
            btn_pick.config(state='disabled')
            btn_config.config(state='disabled')
            root.update_idletasks()

            # progress bar reset
            progress_bar['maximum'] = 100
            progress_bar['value'] = 0

            def progress_cb(current, total):
                try:
                    val = int((current / total) * 100)
                except Exception:
                    val = 0
                progress_bar['value'] = val
                root.update_idletasks()

            df = merge_files(filepaths, header_row=header_row, include_source=include_source, include_sheet=include_sheet, sheet_selection=sheet_selection, progress_callback=progress_cb)
        except Exception as e:
            messagebox.showerror("Error", str(e))
            # re-enable
            btn_merge.config(state='normal')
            btn_pick.config(state='normal')
            btn_config.config(state='normal')
            return
        if df.empty:
            messagebox.showinfo("Info", "Hasil penggabungan kosong.")
            btn_merge.config(state='normal')
            btn_pick.config(state='normal')
            btn_config.config(state='normal')
            return
        # Ask for output file based on selected format
        chosen_fmt = format_var.get().lower()
        if chosen_fmt == 'ndjson':
            filetypes = [("NDJSON", "*.ndjson"), ("JSON", "*.json"), ("All files", "*")]
            default_ext = ".ndjson"
            init_name = "merged.ndjson"
        elif chosen_fmt.startswith('parquet'):
            filetypes = [("Parquet files", "*.parquet"), ("All files", "*")]
            default_ext = ".parquet"
            init_name = "merged.parquet"
        elif chosen_fmt == 'feather':
            filetypes = [("Feather files", "*.feather"), ("All files", "*")]
            default_ext = ".feather"
            init_name = "merged.feather"
        elif chosen_fmt == 'csv-gzip':
            filetypes = [("CSV GZIP", "*.csv.gz"), ("CSV", "*.csv"), ("All files", "*")]
            default_ext = ".csv.gz"
            init_name = "merged.csv.gz"
        elif chosen_fmt == 'csv':
            filetypes = [("CSV", "*.csv"), ("All files", "*")]
            default_ext = ".csv"
            init_name = "merged.csv"
        elif chosen_fmt == 'xlsx':
            filetypes = [("Excel files", "*.xlsx"), ("All files", "*")]
            default_ext = ".xlsx"
            init_name = "merged.xlsx"
        else:
            filetypes = [("JSON files", "*.json *.ndjson"), ("All files", "*")]
            default_ext = ".json"
            init_name = "merged.json"

        out = filedialog.asksaveasfilename(defaultextension=default_ext, filetypes=filetypes, initialfile=init_name)
        if not out:
            # re-enable
            btn_merge.config(state='normal')
            btn_pick.config(state='normal')
            btn_config.config(state='normal')
            return
        try:
            save_dataframe(df, out, fmt=chosen_fmt)
            messagebox.showinfo("Sukses", f"File tersimpan: {out}")
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            btn_merge.config(state='normal')
            btn_pick.config(state='normal')
            btn_config.config(state='normal')
            progress_bar['value'] = 0

    frm = ttk.Frame(root, padding=10)
    frm.pack(fill=tk.BOTH, expand=True)

    lbl = ttk.Label(frm, text="Files:")
    lbl.grid(row=0, column=0, sticky=tk.W)
    txt = tk.Text(frm, height=8, width=80)
    txt.grid(row=1, column=0, columnspan=4, pady=4)

    def refresh_text():
        txt.delete("1.0", tk.END)
        txt.insert(tk.END, files_var.get())

    btn_pick = ttk.Button(frm, text="Pilih Files...", command=lambda: [pick_files(), refresh_text()])
    btn_pick.grid(row=2, column=0, sticky=tk.W)

    btn_config = ttk.Button(frm, text="Konfigurasi Sheet per-file...", command=lambda: configure_sheets_dialog(root, filepaths_getter, sheet_selection))
    btn_config.grid(row=2, column=1, sticky=tk.W, padx=6)

    def filepaths_getter():
        raw = files_var.get().strip()
        return raw.split('\n') if raw else []

    ttk.Label(frm, text="Header row (1-based):").grid(row=3, column=0, sticky=tk.W, pady=6)
    header_entry = ttk.Entry(frm, width=6)
    header_entry.insert(0, "1")
    header_entry.grid(row=3, column=1, sticky=tk.W)

    source_var = tk.BooleanVar(value=True)
    sheet_var = tk.BooleanVar(value=True)
    ttk.Checkbutton(frm, text="Tambah kolom Sumber File", variable=source_var).grid(row=4, column=0, sticky=tk.W)
    ttk.Checkbutton(frm, text="Tambah kolom Sumber Sheet (Excel)", variable=sheet_var).grid(row=4, column=1, sticky=tk.W)
    ttk.Label(frm, text="Format output:").grid(row=4, column=2, sticky=tk.W)
    fmt_combo = ttk.Combobox(frm, textvariable=format_var, values=["json", "ndjson", "parquet-snappy", "parquet-gzip", "feather", "csv-gzip", "csv", "xlsx"], width=18)
    fmt_combo.grid(row=4, column=3, sticky=tk.W)

    btn_merge = ttk.Button(frm, text="Gabung -> Save", command=do_merge)
    btn_merge.grid(row=5, column=0, pady=12, sticky=tk.W)

    # Progress bar
    progress_bar = ttk.Progressbar(frm, orient='horizontal', length=420, mode='determinate')
    progress_bar.grid(row=5, column=1, columnspan=3, padx=6)

    ttk.Label(frm, text="(GUI) Pilih files lalu klik Merge -> Save").grid(row=6, column=0, columnspan=3, sticky=tk.W)

    root.mainloop()


# --------- CLI ----------

def run_cli(argv):
    p = argparse.ArgumentParser(description="Merge files (.txt/.csv/.xls/.xlsx/.parquet/.feather) to a single output file")
    p.add_argument('files', nargs='+', help='Input files (support xlsx, xls, csv, txt, parquet, feather)')
    p.add_argument('-H', '--header', type=int, default=1, help='Header row (1-based)')
    p.add_argument('--no-source', dest='source', action='store_false', help='Do not include Sumber File column')
    p.add_argument('--no-sheet', dest='sheet', action='store_false', help='Do not include Sumber Sheet column')
    p.add_argument('-o', '--output', default='merged.json')
    p.add_argument('--ndjson', action='store_true', help='Save output as NDJSON (newline-delimited)')
    p.add_argument('--format', choices=['json', 'ndjson', 'parquet-snappy', 'parquet-gzip', 'feather', 'csv-gzip', 'csv', 'xlsx'], default='json', help='Output format')
    p.add_argument('--sheet-selection', action='append', help='Specify sheets per file: "path/to/file.xlsx:Sheet1,Sheet2". Repeatable.')
    args = p.parse_args(argv)
    try:
        # parse sheet selections
        sheet_sel = {}
        if getattr(args, 'sheet_selection', None):
            for item in args.sheet_selection:
                if ':' in item:
                    fp, sheets = item.split(':', 1)
                    fp = fp.strip()
                    sheet_list = [s.strip() for s in sheets.split(',') if s.strip()]
                    if sheet_list:
                        sheet_sel[fp] = sheet_list
        outpath = args.output
        # determine chosen format: CLI --ndjson has precedence for backwards compatibility
        chosen_fmt = args.format
        if args.ndjson:
            chosen_fmt = 'ndjson'
        # normalize output extension for some formats
        if chosen_fmt.startswith('parquet') and not outpath.lower().endswith('.parquet'):
            outpath = os.path.splitext(outpath)[0] + '.parquet'
        if chosen_fmt == 'feather' and not outpath.lower().endswith('.feather'):
            outpath = os.path.splitext(outpath)[0] + '.feather'
        if chosen_fmt == 'ndjson' and not outpath.lower().endswith('.ndjson'):
            outpath = os.path.splitext(outpath)[0] + '.ndjson'
        if chosen_fmt == 'csv-gzip' and not outpath.lower().endswith('.csv.gz'):
            outpath = os.path.splitext(outpath)[0] + '.csv.gz'
        if chosen_fmt == 'csv' and not outpath.lower().endswith('.csv'):
            outpath = os.path.splitext(outpath)[0] + '.csv'

        df = merge_files(args.files, header_row=args.header, include_source=args.source, include_sheet=args.sheet, sheet_selection=sheet_sel)
        if df.empty:
            print('Hasil penggabungan kosong')
            return 1
        save_dataframe(df, outpath, fmt=chosen_fmt)
        print(f'Saved to {outpath} ({chosen_fmt})')
        return 0
    except Exception as e:
        print('Error:', e)
        return 2


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '--cli':
        sys.exit(run_cli(sys.argv[2:]))
    else:
        run_gui()
